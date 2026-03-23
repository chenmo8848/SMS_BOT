# -*- coding: utf-8 -*-
"""SMS Bot v6 — Excel 解析服务（短信装弹 + 数据整理，共享基础解析）"""

import io, re, logging
from datetime import datetime
from typing import Optional
from bot.config import BotConfig
from bot.utils.formatting import (
    parse_phone_from_excel, parse_amount, parse_date_for_sms, normalize_phone,
)

log = logging.getLogger(__name__)


def _load_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl, Font, PatternFill, Alignment
    except ImportError:
        return None, None, None, None


def _read_headers(ws, max_scan: int = 5) -> tuple[list[str], int]:
    """读取表头：返回 (headers, header_row)"""
    header_row = 1
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        if any(v for v in row):
            return [str(v).strip() if v is not None else "" for v in row], header_row
        header_row += 1
    return [], 1


# ═══════════════════════════════════════════════════════
#  短信装弹：Excel → 套模板 → 生成发送任务
# ═══════════════════════════════════════════════════════

class SmsExcelResult:
    """短信装弹的处理结果"""
    def __init__(self):
        self.tasks: list[dict] = []
        self.errors: list[str] = []
        self.count: int = 0
        self.amounts: list[float] = []
        self.preview_lines: list[str] = []
        self.txt_content: str = ""


def parse_excel_for_sms(data: bytes, template: str, date_sep: str) -> SmsExcelResult:
    """
    解析 Excel 生成短信任务
    固定五列：姓名、手机号码、银行卡号、放款日期、放款金额
    """
    openpyxl, *_ = _load_openpyxl()
    if not openpyxl:
        r = SmsExcelResult()
        r.errors.append("缺少依赖：openpyxl")
        return r

    result = SmsExcelResult()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers, header_row = _read_headers(ws)

    # 精准列名匹配
    COLS = {
        "name": ["姓名"],
        "phone": ["手机号码", "手机号"],
        "card": ["银行卡号"],
        "date": ["放款日期"],
        "amount": ["放款金额"],
    }

    def find_col(field):
        for name in COLS[field]:
            if name in headers:
                return headers.index(name)
        return -1

    ci = {k: find_col(k) for k in COLS}
    missing = [COLS[k][0] for k, v in ci.items() if v < 0]
    if missing:
        result.errors.append(f"找不到以下列：{'、'.join(missing)}\n识别到的列名：{'、'.join(h for h in headers if h)}")
        return result

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not any(v for v in row):
            continue

        name = str(row[ci["name"]] or "").strip()
        phone = parse_phone_from_excel(row[ci["phone"]])
        card = str(row[ci["card"]] or "").strip().replace(" ", "").replace("-", "")
        date_v = row[ci["date"]]
        amount_v = row[ci["amount"]]

        if not name or not phone or not card:
            result.errors.append(f"第{row_idx}行数据不完整（姓名/手机/卡号）")
            continue
        if len(phone) < 8:
            result.errors.append(f"第{row_idx}行手机号异常：{phone}")
            continue

        date_str = parse_date_for_sms(date_v, sep=date_sep)
        amount_str = parse_amount(amount_v)
        card_last4 = card[-4:] if len(card) >= 4 else card

        msg = (template
               .replace("{姓名}", name)
               .replace("{卡号}", card_last4)
               .replace("{日期}", date_str)
               .replace("{金额}", amount_str))

        result.tasks.append({"phone": phone, "message": msg})
        result.count += 1

        # 金额统计
        try:
            result.amounts.append(float(str(amount_v).replace(",", "")))
        except Exception:
            pass

        # 预览（前3条）
        if len(result.preview_lines) < 3:
            result.preview_lines.append(
                f"👤 {name}  📞 {phone}\n💰 {amount_str} 斤  📅 {date_str}"
            )

    # 生成 txt
    result.txt_content = "\n".join(
        f"{t['phone']}|{t['message']}" for t in result.tasks
    )
    return result


# ═══════════════════════════════════════════════════════
#  数据整理：Excel → 按列配置整理 → 导出
# ═══════════════════════════════════════════════════════

class UserDataResult:
    """数据整理的处理结果"""
    def __init__(self):
        self.xlsx_bytes: Optional[bytes] = None
        self.txt_bytes: Optional[bytes] = None
        self.count: int = 0
        self.skip_rows: list[str] = []
        self.has_time_data: bool = False
        self.error: Optional[str] = None


def _fmt_date_for_user(val, user_date_fmt: str) -> str:
    """智能日期格式化：检测数据精度，按实际精度输出"""
    fmt_has_sec = "%S" in user_date_fmt
    fmt_has_min = "%M" in user_date_fmt
    fmt_has_hour = "%H" in user_date_fmt
    fmt_has_time = fmt_has_hour or fmt_has_min

    def strip_time(fmt):
        fmt = re.sub(r"[ T]?%H:%M:%S", "", fmt)
        fmt = re.sub(r"[ T]?%H:%M", "", fmt)
        fmt = re.sub(r"[ T]?%H", "", fmt)
        return fmt.strip().rstrip("-:/").strip() or "%Y-%m-%d"

    def apply(dt_obj, has_sec, has_min_only):
        if fmt_has_time:
            if has_sec and fmt_has_sec:
                return dt_obj.strftime(user_date_fmt)
            elif (has_sec or has_min_only) and fmt_has_min:
                return dt_obj.strftime(re.sub(r":%S", "", user_date_fmt))
            elif has_min_only and fmt_has_hour:
                return dt_obj.strftime(user_date_fmt)
        return dt_obj.strftime(strip_time(user_date_fmt))

    if hasattr(val, "strftime"):
        has_sec = val.second != 0
        has_min_only = (val.hour != 0 or val.minute != 0) and not has_sec
        has_time = val.hour != 0 or val.minute != 0 or val.second != 0
        if fmt_has_time and has_time:
            return apply(val, has_sec, has_min_only)
        return val.strftime(strip_time(user_date_fmt))

    s = str(val).strip()
    # 含时分秒
    m = re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})[T ](\d{1,2}):(\d{1,2})(?::(\d{1,2}))?", s)
    if m:
        try:
            sec = int(m.group(6)) if m.group(6) else 0
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                          int(m.group(4)), int(m.group(5)), sec)
            data_has_sec = bool(m.group(6))
            data_has_min = True
            if fmt_has_time and (data_has_min or data_has_sec):
                return apply(dt, data_has_sec, data_has_min and not data_has_sec)
            return dt.strftime(strip_time(user_date_fmt))
        except Exception:
            pass

    # 纯日期
    m = re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m and len(s) <= 12:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return dt.strftime(strip_time(user_date_fmt))
        except Exception:
            pass
    if re.match(r"^\d{8}$", s):
        try:
            dt = datetime(int(s[:4]), int(s[4:6]), int(s[6:8]))
            return dt.strftime(strip_time(user_date_fmt))
        except Exception:
            pass
    return s


def _detect_time_data(ws, header_row, date_col_indices, max_rows=10) -> bool:
    """预扫描：检测日期列是否含时分秒"""
    for row in ws.iter_rows(min_row=header_row + 1,
                            max_row=min(header_row + max_rows, ws.max_row),
                            values_only=True):
        for idx in date_col_indices:
            v = row[idx]
            if v is None:
                continue
            if hasattr(v, "strftime"):
                if v.hour != 0 or v.minute != 0 or v.second != 0:
                    return True
            else:
                if re.search(r"\d{1,2}:\d{2}", str(v)):
                    return True
    return False


def parse_excel_for_user(data: bytes, cols: list[str],
                         date_fmt: str) -> UserDataResult:
    """
    按列配置整理 Excel，返回处理结果
    cols: 列名列表，顺序即输出顺序
    date_fmt: 日期格式字符串
    """
    openpyxl, Font, PatternFill, Alignment = _load_openpyxl()
    if not openpyxl:
        r = UserDataResult()
        r.error = "缺少依赖：openpyxl"
        return r

    result = UserDataResult()
    wb_in = openpyxl.load_workbook(io.BytesIO(data))
    ws_in = wb_in.active
    headers, header_row = _read_headers(ws_in)

    # 列匹配
    col_map = {}
    missing = []
    for col_name in cols:
        if col_name in headers:
            col_map[col_name] = headers.index(col_name)
        else:
            missing.append(col_name)

    if missing:
        result.error = (
            f"找不到以下列：{'、'.join(missing)}\n"
            f"Excel 实际列名：{'、'.join(h for h in headers if h)}"
        )
        return result

    # 检测时分秒
    date_cols = [col_map[c] for c in cols
                 if any(kw in c for kw in ["日期", "时间", "date", "Date"])]
    if date_cols:
        result.has_time_data = _detect_time_data(ws_in, header_row, date_cols)

    # 构建输出表
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "用户数据"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for ci, col_name in enumerate(cols, 1):
        cell = ws_out.cell(row=1, column=ci, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 写入数据
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=header_row + 1, values_only=True),
                                  header_row + 1):
        if not any(v for v in row):
            continue
        missing_cols = []
        for col_name in cols:
            idx = col_map[col_name]
            val = row[idx]
            if val is None or str(val).strip() == "":
                missing_cols.append(col_name)
        if missing_cols:
            result.skip_rows.append(f"第{row_idx}行 — 缺少：{'、'.join(missing_cols)}")
            continue

        out_row = []
        for col_name in cols:
            idx = col_map[col_name]
            val = row[idx]
            if any(kw in col_name for kw in ["日期", "时间", "date", "Date"]) and val is not None:
                val = _fmt_date_for_user(val, date_fmt)
            out_row.append(val)
        ws_out.append(out_row)
        result.count += 1

    # 自动列宽
    for col in ws_out.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if result.count == 0:
        if not result.error:
            result.error = "没有有效数据"
        return result

    # 导出 xlsx
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    result.xlsx_bytes = buf.getvalue()

    # 导出 txt（制表符分隔）
    txt_lines = ["\t".join(str(v) for v in cols)]
    for row_data in ws_out.iter_rows(min_row=2, values_only=True):
        txt_lines.append("\t".join(str(v or "") for v in row_data))
    result.txt_bytes = "\n".join(txt_lines).encode("utf-8-sig")

    return result


# ═══════════════════════════════════════════════════════
#  批量文本解析：txt/粘贴文本 → 任务列表
# ═══════════════════════════════════════════════════════

def parse_batch_text(text: str) -> tuple[list[dict], list[str]]:
    """解析 手机号|内容 格式的文本，返回 (tasks, errors)"""
    tasks = []
    errors = []
    for i, line in enumerate(text.strip().splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        if "|" not in line:
            errors.append(f"第{i}行格式错误（缺少 | 分隔符）")
            continue
        p, m = line.split("|", 1)
        p, m = p.strip(), m.strip()
        if not p or not m:
            errors.append(f"第{i}行号码或内容为空")
            continue
        tasks.append({"phone": p, "message": m})
    return tasks, errors


def parse_batch_file(data: bytes) -> tuple[Optional[str], str]:
    """
    尝试多种编码解码文件内容
    返回 (decoded_text, error_msg)
    """
    for enc in ["utf-8-sig", "utf-8", "gbk", "gb2312", "latin1"]:
        try:
            return data.decode(enc), ""
        except (UnicodeDecodeError, LookupError):
            continue
    return None, "文件编码无法识别，请使用 UTF-8 编码"
